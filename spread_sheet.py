
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

def write_sheet(USN, Name, marks, ec, tc, sgpa):
    wb = Workbook()
    sheet = wb.active
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    sheet.cell(row=1, column=1).value = "USN"
    sheet.cell(row=1, column=2).value = USN
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
    sheet.cell(row=2, column=1).value = "Name"
    sheet.cell(row=2, column=2).value = Name
    head = ['SUB-CODE', 'SUBJECT-NAME', 'TOTAL-CREDITS',
            'TOTAL-MARKS', 'GRADE-LETTER', 'GRADE-POINT', 'EARNED-CREDITS']
    for i in range(0, len(head)):
        sheet.column_dimensions[get_column_letter(i+1)].width = 20
        sheet.cell(row=3, column=i+1).font = Font(bold=True, color="0099CCFF")
        sheet.cell(row=3, column=i+1).value = head[i]
    for i in marks:
        sheet.append(i)
    r = sheet.max_row+1

    sheet.append(["Earned Credits ", ec])
    sheet.append(["Total Credits ", tc])
    sheet.append(["SGPA", sgpa])
    sheet.cell(row=r, column=1).font = Font(bold=True, color="0099CCFF")
    sheet.cell(row=r+1, column=1).font = Font(bold=True, color="0099CCFF")
    sheet.cell(row=r+2, column=1).font = Font(bold=True, color="0099CCFF")

    wb.save(f"{USN}.xlsx")
    # file = f'C:/Users/ROHAN/mypython/res-scrape/res_scrape/{USN}.xlsx'
    # os.startfile(file)

#                           Test Data
#     marks = [['18CS51', 'MANAGEMENT AND ENTREPRENEURSHIP FOR IT INDUSTRY', 3, 83, 'S', 9, 27],
# ['18CS52', 'COMPUTER NETWORKS AND SECURITY', 4, 86, 'S', 9, 36],
#  ['18CS53', 'DATABASE MANAGEMENT SYSTEMS', 4, 79, 'A', 8, 32],
#  ['18CS54', 'AUTOMATA THEORY AND COMPUTABILITY', 3, 62, 'B', 7, 21],
#   ['18CS55', 'APPLICATION DEVELOPMENT USING PYTHON', 3, 78, 'A', 8, 24],
#    ['18CS56', 'UNIX PROGRAMMING', 3, 75, 'A', 8, 24],
#    ['18CSL57', 'COMPUTER NETWORK LABORATORY', 2, 90, 'O', 10, 20],
#    ['18CSL58', 'DBMS LABORATORY WITH MINI PROJECT', 2, 88, 'S', 9, 18],
#    ['18CIV59', 'ENVIRONMENTAL STUDIES', 1, 76, 'A', 8, 8]]
    # ec = 240
    # tc = 25
    # sgpa = 7.08
